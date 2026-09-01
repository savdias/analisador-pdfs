dicionario_de_busca = {
    "Antigas": [
        "agronegócio", "negócio agrícola", #... etc
    ]
}
import os
import re
import fitz  # PyMuPDF
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from concurrent.futures import ProcessPoolExecutor

# =========================================================================
# FUNÇÃO DE DETECÇÃO DE GÊNERO TEXTUAL POR TRECHO / PARÁGRAFO (AMPLIADA)
# =========================================================================
def identificar_genero_trecho(texto_trecho):
    """
    Analisa as marcas linguísticas, estruturais e lexicais de um trecho/parágrafo específico
    para classificar o seu gênero textual dentre uma lista abrangente.
    """
    texto = texto_trecho.lower()

    pontuacao = {
        'Poema / Poesia / Cordel': 0,
        'História em Quadrinhos / Charge / Cartum / Tirinha': 0,
        'Narrativa Literária (Conto / Crônica / Fábula / Mito)': 0,
        'Peça Teatral / Roteiro Dramático': 0,
        'Biografia / Autobiografia': 0,
        'Notícia / Reportagem / Texto Jornalístico': 0,
        'Artigo de Opinião / Editorial / Ensaio / Manifesto': 0,
        'Resenha / Crítica Literária': 0,
        'Entrevista / Depoimento': 0,
        'Anúncio Publicitário / Propaganda / Campanha': 0,
        'Receita Culinária': 0,
        'Manual Técnico / Tutorial / Regra de Jogo': 0,
        'Artigo Científico / Acadêmico / Relatório': 0,
        'Verbete de Enciclopédia / Dicionário / Glossário': 0,
        'Carta / E-mail / Correspondência': 0,
        'Atividade / Exercício Didático': 0,
        'Texto Legislativo / Documento Oficial': 0,
        'Infográfico / Legenda / Descrição de Tabela': 0,
        'Discurso / Pronunciamento': 0,
        'Texto Expositivo / Didático Geral': 0
    }

    # 1. Poema / Poesia / Cordel
    if re.search(r'\b(estrofe|estrofes|verso|versos|rima|rimas|poema|poesia|cordel|trova|soneto|lírico|trovadorismo)\b', texto):
        pontuacao['Poema / Poesia / Cordel'] += 3
    if texto.count('/') >= 2 or texto.count('\n') >= 2:
        linhas = [l.strip() for l in re.split(r'[\n/]', texto_trecho) if l.strip()]
        if len(linhas) >= 3 and (sum(len(l.split()) for l in linhas) / len(linhas)) < 9:
            pontuacao['Poema / Poesia / Cordel'] += 4

    # 2. História em Quadrinhos / Charge / Cartum / Tirinha
    if re.search(r'\b(quadrinho|quadrinhos|hq|charge|cartum|tirinha|tirinhas|balão de fala|balões|quadro \d+|tira \d+)\b', texto):
        pontuacao['História em Quadrinhos / Charge / Cartum / Tirinha'] += 4

    # 3. Peça Teatral / Roteiro Dramático
    if re.search(r'\b(cena \d+|ato \d+|rubrica|cenário|personagens:|entrando em cena|sai de cena)\b', texto) or re.search(r'^[A-ZÁÉÍÓÚÂÊÔÃÕÇ]{3,}\s*\(.*\)\s*:', texto_trecho, re.MULTILINE):
        pontuacao['Peça Teatral / Roteiro Dramático'] += 4

    # 4. Biografia / Autobiografia
    if re.search(r'\b(nasceu em|faleceu em|morreu em|sua infância|sua vida|trajetória|biografia|autobiografia|formou-se|obras principais|escritor brasileiro|poeta nascido|vida e obra)\b', texto):
        pontuacao['Biografia / Autobiografia'] += 4

    # 5. Narrativa Literária (Conto, Crônica, Fábula, Mito, Lenda)
    if re.search(r'\b(era uma vez|certo dia|de repente|personagem|narrador|disse|respondeu|perguntou|exclamou|havia um|moral da história|numa noite|fábula|mito|lenda)\b', texto) or '—' in texto_trecho or '–' in texto_trecho:
        pontuacao['Narrativa Literária (Conto / Crônica / Fábula / Mito)'] += 2
        if re.search(r'\b(moral da história|fábula|era uma vez)\b', texto):
            pontuacao['Narrativa Literária (Conto / Crônica / Fábula / Mito)'] += 3

    # 6. Notícia / Reportagem / Texto Jornalístico
    if re.search(r'\b(reportagem|notícia|jornal|texto jornalístico|manchete|jornalista|redação|segundo o órgão|divulgado|segundo a polícia|ocorreu na|afirmou o especialista|foto:|fonte:)\b', texto):
        pontuacao['Notícia / Reportagem / Texto Jornalístico'] += 3
    if re.search(r'\b(segundo|de acordo com|nesta (segunda|terça|quarta|quinta|sexta)|ontem|na manhã de|governo estadual|prefeitura|agência de notícias)\b', texto):
        pontuacao['Notícia / Reportagem / Texto Jornalístico'] += 1

    # 7. Artigo de Opinião / Editorial / Ensaio / Manifesto
    if re.search(r'\b(em minha opinião|na minha visão|defendemos que|acreditamos|é preciso considerar|lamentavelmente|urgente|em suma|portanto|convém destacar|ensaio|artigo de opinião|editorial|manifesto)\b', texto):
        pontuacao['Artigo de Opinião / Editorial / Ensaio / Manifesto'] += 3

    # 8. Resenha / Crítica Literária
    if re.search(r'\b(resenha|crítica|obra analisada|livro de|dirigido por|recomenda-se|pontos fortes|análise do filme|livro traz|leitura indispensável)\b', texto):
        pontuacao['Resenha / Crítica Literária'] += 3

    # 9. Entrevista / Depoimento
    if re.search(r'\b(entrevistado|entrevistador|pergunta:|resposta:|depoimento|afirma o especialista|entrevista concedida|relata que|entrevistado:)\b', texto):
        pontuacao['Entrevista / Depoimento'] += 4

    # 10. Anúncio Publicitário / Propaganda / Campanha
    if re.search(r'\b(compre|aproveite|oferta|desconto|ligue já|promoção|garanta|não perca|confira|slogan|propaganda|anúncio|campanha|vacine-se)\b', texto):
        pontuacao['Anúncio Publicitário / Propaganda / Campanha'] += 3

    # 11. Receita Culinária
    if re.search(r'\b(ingredientes|modo de preparo|rendimento|modo de fazer|colher de|xícara|pitada|fogo médio|forno preaquecido|sirva em seguida)\b', texto):
        pontuacao['Receita Culinária'] += 5

    # 12. Manual Técnico / Tutorial / Regra de Jogo
    if re.search(r'\b(passo \d|passo a passo|instruções|manual|configuração|procedimento|instalação|modo de usar|requisitos|clique em|conecte|regras do jogo|como jogar)\b', texto):
        pontuacao['Manual Técnico / Tutorial / Regra de Jogo'] += 3

    # 13. Artigo Científico / Acadêmico / Relatório
    if re.search(r'\b(resumo|abstract|metodologia|referências bibliográficas|hipótese|metodológico|amostra|conclui-se|introdução|fundamentação teórica|relatório de experimento)\b', texto):
        pontuacao['Artigo Científico / Acadêmico / Relatório'] += 3

    # 14. Verbete de Enciclopédia / Dicionário / Glossário
    if re.search(r'\b(substantivo|adjetivo|verbo|s\.f\.|s\.m\.|etimologia|definido como|refere-se a|termo utilizado para|enciclopédia|glossário)\b', texto):
        pontuacao['Verbete de Enciclopédia / Dicionário / Glossário'] += 3

    # 15. Carta / E-mail / Correspondência
    if re.search(r'\b(prezado|prezada|caro|cara|atenciosamente|cordialmente|remetente|destinatário|assunto:|anexo|abraços|carta aberta|estimado)\b', texto):
        pontuacao['Carta / E-mail / Correspondência'] += 3

    # 16. Atividade / Exercício Didático
    if re.search(r'\b(exercício|exercícios|atividade|questão|questões|responda|assinale|marque|leia o texto|observe a figura|veja a imagem|saiba mais|analise o gráfico|com base no texto|faça o que se pede|resolva|assinale a alternativa)\b', texto):
        pontuacao['Atividade / Exercício Didático'] += 3

    # 17. Texto Legislativo / Documento Oficial
    if re.search(r'\b(artigo \d+|art\. \d+|parágrafo único|inciso|decreto|lei nº|constituição|portaria|edital|resolução)\b', texto):
        pontuacao['Texto Legislativo / Documento Oficial'] += 4

    # 18. Infográfico / Legenda / Descrição de Tabela
    if re.search(r'\b(fonte:|gráfico \d|figura \d|tabela \d|organograma|infográfico|legenda:)\b', texto):
        pontuacao['Infográfico / Legenda / Descrição de Tabela'] += 2

    # 19. Discurso / Pronunciamento
    if re.search(r'\b(senhores e senhoras|caros cidadãos|caros colegas|discurso proferido|pronunciamento|em meu nome e em nome)\b', texto):
        pontuacao['Discurso / Pronunciamento'] += 3

    # 20. Texto Expositivo / Didático Geral
    if re.search(r'\b(conceito|definição|podemos definir|consiste em|caracteriza-se por|exemplo|é chamado de|estudo de|área de|denomina-se)\b', texto):
        pontuacao['Texto Expositivo / Didático Geral'] += 1

    genero_predominante = max(pontuacao, key=pontuacao.get)

    if pontuacao[genero_predominante] == 0:
        return "Texto Expositivo / Didático Geral"

    return genero_predominante

# =========================================================================
# FUNÇÃO DE LIMPEZA GERAL
# =========================================================================
def limpar_caracteres_invalidos(texto):
    if not isinstance(texto, str):
        return texto
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', texto)

# =========================================================================
# WORKER PARALELO OTIMIZADO
# =========================================================================
def _processar_lote_paginas_texto(argumentos):
    caminho_pdf, lote_paginas, mapa_palavras = argumentos
    doc = fitz.open(caminho_pdf)
    resultados_lote = []
    
    padroes_compilados = {}
    for palavra, categoria in mapa_palavras.items():
        padroes_compilados[palavra] = (
            categoria, 
            re.compile(r'(.{0,150}\b' + re.escape(palavra) + r'\b.{0,150})', flags=re.IGNORECASE)
        )
        
    for num_pagina in lote_paginas:
        pagina = doc[num_pagina]
        texto_pagina = pagina.get_text()
        
        if not texto_pagina: 
            continue
            
        texto_pagina_limpo = texto_pagina.replace('\n', ' ')
        texto_pagina_lower = texto_pagina_limpo.lower()
        
        for palavra, (categoria, padrao) in padroes_compilados.items():
            if palavra.lower() not in texto_pagina_lower: 
                continue
                
            for ocorrencia in padrao.finditer(texto_pagina_limpo):
                trecho = ocorrencia.group(0).strip()
                
                # Classificação do gênero feita diretamente no trecho/parágrafo
                genero_trecho = identificar_genero_trecho(trecho)

                resultados_lote.append({
                    'Categoria': categoria,
                    'Palavra-Chave': palavra,
                    'Página': num_pagina + 1,
                    'Gênero Textual do Trecho': genero_trecho,
                    'Trecho Encontrado': f"... {trecho} ..."
                })
                
    doc.close()
    return resultados_lote

# =========================================================================
# FUNÇÃO PRINCIPAL DE EXTRAÇÃO E FORMATAÇÃO
# =========================================================================
def analisar_pdf_para_planilha(caminho_pdf, dicionario_palavras, caminho_saida_excel):
    print(f"Iniciando a leitura do arquivo: {caminho_pdf}")
    
    doc = fitz.open(caminho_pdf)
    total_paginas = len(doc)
    doc.close()

    mapa_palavras = {}
    vistos = set()
    for categoria, palavras in dicionario_palavras.items():
        for p in palavras:
            p_clean = p.strip()
            p_lower = p_clean.lower()
            if p_lower not in vistos:
                vistos.add(p_lower)
                mapa_palavras[p_clean] = categoria

    tamanho_lote = 30
    lotes_paginas = [list(range(i, min(i + tamanho_lote, total_paginas))) for i in range(0, total_paginas, tamanho_lote)]
    tarefas = [(caminho_pdf, lote, mapa_palavras) for lote in lotes_paginas]
    resultados = []
    
    with ProcessPoolExecutor() as executor:
        listas_de_resultados = executor.map(_processar_lote_paginas_texto, tarefas)
        for lista in listas_de_resultados:
            resultados.extend(lista)
            
    df_base = pd.DataFrame(resultados)

    if df_base.empty:
        print("Nenhuma das palavras-chave foi encontrada.")
        return

    df_base['Trecho Encontrado'] = df_base['Trecho Encontrado'].apply(limpar_caracteres_invalidos)
    df_base = df_base.drop_duplicates(subset=['Categoria', 'Palavra-Chave', 'Página', 'Trecho Encontrado']).reset_index(drop=True)
    
    # 1. ABA DE QUANTIFICAÇÃO DE PALAVRAS
    df_quantificacao = df_base.groupby(['Categoria', 'Palavra-Chave']).size().reset_index(name='Quantidade de Aparições')
    df_quantificacao = df_quantificacao.sort_values(by=['Categoria', 'Palavra-Chave'])
    
    # 2. ABA DE GÊNEROS TEXTUAIS (Com páginas e ocorrências)
    df_generos = df_base.groupby('Gênero Textual do Trecho').agg({
        'Gênero Textual do Trecho': 'count',
        'Página': lambda x: ', '.join(map(str, sorted(set(x))))
    }).rename(columns={
        'Gênero Textual do Trecho': 'Quantidade de Ocorrências', 
        'Página': 'Páginas Encontradas'
    }).reset_index()
    
    df_paginas = df_base.groupby(['Categoria', 'Palavra-Chave'])['Página'].apply(lambda x: ', '.join(map(str, sorted(set(x))))).reset_index(name='Páginas Encontradas')
    df_paginas = df_paginas.sort_values(by=['Categoria', 'Palavra-Chave'])
    
    df_todas_paginas = pd.DataFrame({'Páginas com Ocorrências (Visão Geral)': sorted(df_base['Página'].unique())})

    # Bloco formatado exibindo o Gênero Textual do próprio trecho
    df_agrupado_palavra = df_base.groupby(['Categoria', 'Palavra-Chave', 'Página', 'Gênero Textual do Trecho'])['Trecho Encontrado'].apply(lambda trechos: '\n\n'.join(trechos)).reset_index()
    df_agrupado_palavra['Bloco Formatado'] = df_agrupado_palavra.apply(
        lambda row: f"► {row['Palavra-Chave']} ({row['Categoria']} | Gênero: {row['Gênero Textual do Trecho']}):\n{row['Trecho Encontrado']}", axis=1
    )
    
    # Agrupamento por Página
    df_aba1 = df_agrupado_palavra.groupby('Página').agg({
        'Categoria': lambda x: ', '.join(sorted(set(x))),
        'Palavra-Chave': lambda x: ', '.join(sorted(set(x))),
        'Gênero Textual do Trecho': lambda x: ', '.join(sorted(set(x))),
        'Bloco Formatado': lambda x: '\n\n--------------------------------------------------\n\n'.join(x)
    }).reset_index()
    
    df_aba1.rename(columns={
        'Bloco Formatado': 'Trechos Encontrados', 
        'Categoria': 'Categorias na Página', 
        'Palavra-Chave': 'Palavras-Chave',
        'Gênero Textual do Trecho': 'Gêneros Identificados na Página'
    }, inplace=True)
    
    df_aba1 = df_aba1[['Página', 'Categorias na Página', 'Palavras-Chave', 'Gêneros Identificados na Página', 'Trechos Encontrados']]

    # Salvando no Excel e aplicando formatação
    with pd.ExcelWriter(caminho_saida_excel, engine='openpyxl') as writer:
        df_aba1.to_excel(writer, index=False, sheet_name='Análise Detalhada')
        df_quantificacao.to_excel(writer, index=False, sheet_name='Quantificação de Palavras')
        df_generos.to_excel(writer, index=False, sheet_name='Gêneros Textuais')
        df_paginas.to_excel(writer, index=False, sheet_name='Páginas por Palavra')
        df_todas_paginas.to_excel(writer, index=False, sheet_name='Resumo de Páginas')
        df_base.to_excel(writer, index=False, sheet_name='Base de Dados') 

        fonte_cabecalho = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        fonte_dados = Font(name="Calibri", size=11)

        writer.sheets['Base de Dados'].sheet_state = 'hidden'

        # Estilo da Aba 1 (Análise Detalhada)
        ws1 = writer.sheets['Análise Detalhada']
        cor_azul = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        for cell in ws1[1]:
            cell.fill = cor_azul
            cell.font = fonte_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws1.column_dimensions['A'].width = 12  
        ws1.column_dimensions['B'].width = 25  
        ws1.column_dimensions['C'].width = 30  
        ws1.column_dimensions['D'].width = 35  
        ws1.column_dimensions['E'].width = 100 
        
        for row in ws1.iter_rows(min_row=2, max_col=5, max_row=ws1.max_row):
            for cell in row:
                cell.font = fonte_dados
                cell.alignment = Alignment(
                    vertical="top", 
                    horizontal="center" if cell.column_letter == 'A' else "left",
                    wrap_text=(cell.column_letter in ['B', 'C', 'D', 'E'])
                )
            
        # Estilo da Aba 2 (Quantificação de Palavras)
        ws2 = writer.sheets['Quantificação de Palavras']
        cor_verde = PatternFill(start_color="2E6B4E", end_color="2E6B4E", fill_type="solid")
        for cell in ws2[1]:
            cell.fill = cor_verde
            cell.font = fonte_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws2.column_dimensions['A'].width = 20  
        ws2.column_dimensions['B'].width = 30  
        ws2.column_dimensions['C'].width = 25  
        
        for row in ws2.iter_rows(min_row=2, max_col=3, max_row=ws2.max_row):
            for cell in row:
                cell.font = fonte_dados
                cell.alignment = Alignment(vertical="center", horizontal="left" if cell.column_letter != 'C' else "center")

        # Estilo da Aba 3 (Gêneros Textuais)
        ws_genero = writer.sheets['Gêneros Textuais']
        cor_azul_escuro = PatternFill(start_color="0F243E", end_color="0F243E", fill_type="solid")
        for cell in ws_genero[1]:
            cell.fill = cor_azul_escuro
            cell.font = fonte_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws_genero.column_dimensions['A'].width = 45  
        ws_genero.column_dimensions['B'].width = 25  
        ws_genero.column_dimensions['C'].width = 50  
        
        for row in ws_genero.iter_rows(min_row=2, max_col=3, max_row=ws_genero.max_row):
            for cell in row:
                cell.font = fonte_dados
                cell.alignment = Alignment(vertical="top", horizontal="center" if cell.column_letter == 'B' else "left", wrap_text=True)

        # ESCALA DE CORES NA COLUNA DE QUANTIDADE DE OCORRÊNCIAS (COLUNA B)
        if ws_genero.max_row >= 2:
            regra_escala_cores = ColorScaleRule(
                start_type='min', start_color='E0F2FE',  # Azul suave (menor volume)
                mid_type='percentile', mid_value=50, mid_color='7DD3FC', # Azul médio (volume mediano)
                end_type='max', end_color='0284C7'       # Azul intenso (maior volume)
            )
            ws_genero.conditional_formatting.add(f'B2:B{ws_genero.max_row}', regra_escala_cores)

        # Estilo da Aba 4 (Páginas por Palavra)
        ws3 = writer.sheets['Páginas por Palavra']
        cor_laranja = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
        for cell in ws3[1]:
            cell.fill = cor_laranja
            cell.font = fonte_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws3.column_dimensions['A'].width = 15  
        ws3.column_dimensions['B'].width = 30  
        ws3.column_dimensions['C'].width = 50  
        
        for row in ws3.iter_rows(min_row=2, max_col=3, max_row=ws3.max_row):
            for cell in row:
                cell.font = fonte_dados
                cell.alignment = Alignment(vertical="center", horizontal="left" if cell.column_letter != 'C' else "center", wrap_text=True)
        
        # Estilo da Aba 5 (Resumo de Páginas)
        ws4 = writer.sheets['Resumo de Páginas']
        cor_roxa = PatternFill(start_color="604A7B", end_color="604A7B", fill_type="solid")
        for cell in ws4[1]:
            cell.fill = cor_roxa
            cell.font = fonte_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws4.column_dimensions['A'].width = 40  
        
        for row in ws4.iter_rows(min_row=2, max_col=1, max_row=ws4.max_row):
            for cell in row:
                cell.font = fonte_dados
                cell.alignment = Alignment(vertical="center", horizontal="center")
    
    print(f"Sucesso! Arquivo Excel extraído e salvo em: {caminho_saida_excel}\n")

if __name__ == "__main__":
    dicionario_de_busca = {
        "Antigas": [
            "agronegócio", "negócio agrícola", "negócio agropecuário", "agropecuária",
            "setor agropecuário", "setor agrícola", "setor rural", "cadeia produtiva",
            "cadeia produtiva agropecuária", "cadeia agroindustrial", "complexo agroindustrial",
            "agroindústria", "sistema agroalimentar", "sistema alimentar", "produção rural",
            "produção do campo", "setor primário", "campo e indústria", "campo e cidade",
            "produção e mercado", "produção e consumo", "insumos", "insumos agrícolas",
            "insumos agropecuários", "sementes", "mudas", "genética", "melhoramento genético",
            "fertilizantes", "adubos", "adubação", "corretivos", "defensivos", "defensivos agrícolas",
            "agrotóxicos", "pesticidas", "herbicidas", "inseticidas", "fungicidas", "bioinsumos",
            "máquinas", "máquinas agrícolas", "tratores", "implementos", "irrigação", "crédito rural",
            "financiamento", "seguro rural", "assistência técnica", "extensão rural",
            "planejamento produtivo", "pesquisa agropecuária", "agricultura", "agrícola",
            "lavoura", "plantio", "cultivo", "colheita", "safra", "entressafra", "produtividade",
            "monocultura", "policultura", "agricultura intensiva", "agricultura extensiva",
            "culturas permanentes", "culturas temporárias", "soja", "milho", "café", "algodão",
            "arroz", "feijão", "trigo", "cana-de-açúcar", "cana", "frutas", "fruticultura",
            "hortaliças", "horticultura", "citricultura", "silvicultura", "pecuária",
            "criação animal", "rebanho", "gado", "bovino", "bovinocultura", "suinocultura",
            "avicultura", "caprinocultura", "ovinocultura", "piscicultura", "aquicultura",
            "produção de leite", "produção de carne", "produção de ovos", "confinamento",
            "pastagem", "manejo animal", "sanidade animal", "vacinação animal", "abate",
            "criação intensiva", "criação extensiva", "pecuária intensiva", "pecuária extensiva",
            "beneficiamento", "processamento", "transformação", "industrialização",
            "agroindustrial", "frigorífico", "laticínio", "abatedouro", "usina", "moagem",
            "torrefação", "pasteurização", "conservação", "processamento de alimentos",
            "indústria de alimentos", "indústria alimentícia", "indústria de bebidas",
            "indústria sucroenergética", "etanol", "biodiesel", "biocombustível",
            "biocombustíveis", "celulose", "papel", "óleo vegetal", "farelo", "açúcar",
            "suco", "processamento de carnes", "processamento de leite", "fabricação de derivados",
            "armazenamento", "armazenagem", "silos", "armazéns", "entrepostos", "transporte",
            "rodovias", "ferrovias", "hidrovias", "portos", "logística", "distribuição",
            "comercialização", "atacado", "varejo", "exportação", "importação", "mercado interno",
            "mercado externo", "commodities", "commodity", "preço agrícola", "preços agrícolas",
            "bolsa de mercadorias", "agregação de valor", "rastreabilidade", "cadeia de frio",
            "embalagem", "escoamento da produção", "agricultura de precisão", "mecanização",
            "automação", "sensores", "drones", "satélites", "georreferenciamento", "biotecnologia",
            "transgenia", "transgênico", "organismos geneticamente modificados",
            "inteligência de dados", "monitoramento", "irrigação inteligente", "plantio direto",
            "controle biológico", "agricultura digital", "conectividade rural", "inovação agropecuária",
            "inovação agrícola", "agricultura 4.0", "automação agroindustrial", "desmatamento",
            "erosão", "contaminação", "uso da água", "consumo de água", "uso do solo",
            "degradação", "emissao", "emissões", "gases de efeito estufa", "metano", "biodiversidade",
            "queimadas", "resíduos", "efluentes", "recuperação", "mitigação", "preservação",
            "sustentabilidade", "sustentável", "economia circular", "reaproveitamento",
            "redução de perdas", "aproveitamento de resíduos", "recuperação de solos",
            "integração lavoura-pecuária-floresta", "ILPF", "mata ciliar", "reserva legal",
            "área de preservação permanente", "APP", "emprego", "renda", "PIB",
            "produto interno bruto", "balança comercial", "trabalho rural", "trabalhador rural",
            "mecanização do trabalho", "êxodo rural", "concentração fundiária", "estrutura fundiária",
            "reforma agrária", "cooperativismo", "cooperativa", "associativismo", "assentamento",
            "conflito no campo", "segurança alimentar", "soberania alimentar", "abastecimento",
            "relações campo-cidade", "propriedade rural", "pequena propriedade", "grande propriedade",
            "latifúndio", "minifúndio", "agricultura familiar", "agricultura empresarial",
            "agricultura comercial", "agricultura de subsistência", "agroecologia",
            "agricultura orgânica", "produção orgânica", "camponês", "campesinato",
            "pequeno produtor", "grande produtor", "agricultura convencional", "produção integrada",
            "sistema agroflorestal", "extrativismo vegetal", "produção sustentável", "política agrícola",
            "política agrária", "crédito agrícola", "regularização fundiária", "zoneamento agrícola",
            "fiscalização", "licenciamento", "legislação ambiental", "código florestal",
            "vigilância sanitária", "inspeção sanitária", "certificação", "subsídio", "incentivo fiscal",
            "seguro agrícola", "direitos trabalhistas", "trabalho análogo à escravidão",
            "trabalho infantil", "conflito fundiário", "fome", "consumo", "alimentação",
            "alimento", "alimentos", "desperdício", "desperdício de alimentos", "perda de alimentos",
            "perdas pós-colheita", "conservação de alimentos", "qualidade dos alimentos",
            "origem dos alimentos", "rastreabilidade dos alimentos", "cadeia alimentar",
            "preço dos alimentos", "acesso aos alimentos"
        ]
    }
    
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    lista_de_livros = [os.path.join(diretorio_script, "pdf", f"livro{i}.pdf") for i in range(19, 20)]
    
    for arquivo_pdf in lista_de_livros:
        if os.path.exists(arquivo_pdf):
            nome_saida = os.path.splitext(os.path.basename(arquivo_pdf))[0] + "_analise.xlsx"
            caminho_saida = os.path.join(diretorio_script, nome_saida)
            analisar_pdf_para_planilha(arquivo_pdf, dicionario_de_busca, caminho_saida)
        else:
            print(f"Arquivo não encontrado: {arquivo_pdf}")